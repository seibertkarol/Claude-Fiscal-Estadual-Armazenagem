"""
CONCILIACAO DE ANULACOES (-002) — SLC Agricola
===============================================

Pareia cada nota de ANULACAO (referencia com sufixo -002 na Planilha6) com a
nota de ORIGEM que ela esta anulando, e pinta os dois lados de ROXO.

Garantia: filtrar a Planilha6 pela cor ROXA deve somar exatamente R$0,00.

Uso:
  py -3.14 concilia_anulacoes.py --arquivo="PLANILHA.xlsx" [--relatorio="EXPORT_ERP.xlsx"]

  --arquivo   : Excel com a aba Planilha6 (crua ou ja conciliada)
  --relatorio : (opcional) relatorio do ERP com as notas de anulacao.
                Sem ele, o script usa so o metodo de valor unico.

DOIS METODOS DE PAREAMENTO
--------------------------
1. ERP (preferencial) — o relatorio traz, para cada anulacao, o documento que
   ela esta anulando:
     coluna A  'N documento'                       -> doc_sap da anulacao
     coluna AF 'Numero de nota fiscal eletronica'  -> NF da anulacao
     coluna BE 'N doc.original'                    -> doc_sap da origem
     coluna BF 'N NOTA ORIGEM'                     -> NF da origem
   No relatorio a anulacao aparece SEM o sufixo -002 (so o numero da nota).

2. VALOR UNICO (fallback) — procura uma linha com o valor exatamente invertido.
   So aceita quando existe UMA UNICA candidata E nenhuma outra anulacao disputa
   o mesmo valor. Valores repetidos (ex: R$35,73 aparece centenas de vezes)
   ficam sem par de proposito, para nao arriscar um falso positivo.

REGRAS QUE GARANTEM O SALDO ZERO
--------------------------------
- O pareamento e por LINHA, nunca por numero de nota. Um mesmo numero pode ter
  varias linhas na planilha e so uma delas e o par.
- Linhas ja ESTORNADAS (X em 'estornado' ou em 'Documento de estorno') nunca
  entram como origem — elas ja foram neutralizadas por outra linha.
- Cada linha de origem so pode ser consumida por UMA anulacao. Se duas
  anulacoes apontam para a mesma origem, a segunda vira CONFLITO e nao e
  pintada (caso real: duas anulacoes para o mesmo doc_sap no ERP).
"""

import pandas as pd
import sys, os, unicodedata, warnings

warnings.filterwarnings('ignore')   # datas em formatos mistos geram aviso a cada linha
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

ROXO_HEX = 'CC99FF'   # cor das anulacoes — nao colide com as cores da conciliacao
                      # de armazenagem (rosa/laranja/azul/amarelo/verde)

# ---------------------------------------------------------------
# PARAMETROS
# ---------------------------------------------------------------
ARQUIVO = RELATORIO = None
for arg in sys.argv[1:]:
    if arg.startswith('--arquivo='):
        ARQUIVO = arg.split('=', 1)[-1].strip().strip('"')
    elif arg.startswith('--relatorio='):
        RELATORIO = arg.split('=', 1)[-1].strip().strip('"')

if not ARQUIVO:
    print('ERRO: informe o Excel com --arquivo="NOME.xlsx"')
    sys.exit(1)
if not os.path.exists(ARQUIVO):
    print(f'ERRO: arquivo nao encontrado: {ARQUIVO}')
    sys.exit(1)
if RELATORIO and not os.path.exists(RELATORIO):
    print(f'ERRO: relatorio nao encontrado: {RELATORIO}')
    sys.exit(1)


def _norm(s):
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode('ascii')
    return s.lower().strip()


def _find_col(df, keyword, fallback_idx, rotulo=''):
    """Localiza coluna pelo nome do cabecalho; cai na posicao fixa se nao achar.

    Tenta primeiro igualdade exata (evita que 'documento' case com
    'doc.original'), depois substring, e por fim procura o cabecalho na
    primeira linha de dados (planilhas cujo cabecalho real nao virou nome
    de coluna do pandas — as colunas saem como 'Unnamed: N').
    """
    kw = _norm(keyword)
    for i, c in enumerate(df.columns):
        if _norm(c) == kw:
            return i
    for i, c in enumerate(df.columns):
        if kw in _norm(c):
            return i
    if len(df) > 0:
        primeira = df.iloc[0].tolist()
        for i, v in enumerate(primeira):
            if _norm(v) == kw:
                return i
        for i, v in enumerate(primeira):
            if kw in _norm(v):
                return i
    print(f"  [{rotulo}] coluna '{keyword}' nao encontrada, usando posicao fixa {fallback_idx}")
    return fallback_idx


# ---------------------------------------------------------------
# LEITURA DA PLANILHA6
# ---------------------------------------------------------------
print('Carregando Planilha6...')
_wb = load_workbook(ARQUIVO, read_only=True, data_only=True)
P6_NOME = next((s for s in _wb.sheetnames if s.lower() == 'planilha6'), 'Planilha6')
JA_CONCILIADO = (_wb[P6_NOME]['A1'].value == 'TOTAL FILTRADO')
_wb.close()

# planilha ja conciliada tem a linha TOTAL FILTRADO no topo e a coluna J
# (DATA DO RETORNO) inserida, o que desloca cabecalho e colunas
HEADER    = 1 if JA_CONCILIADO else 0
LINHA_OFF = 3 if JA_CONCILIADO else 2   # pandas idx -> linha do Excel
COL_OFF   = 1 if JA_CONCILIADO else 0   # deslocamento das colunas a partir de J

df = pd.read_excel(ARQUIVO, sheet_name=P6_NOME, header=HEADER)
print(f"  arquivo {'JA CONCILIADO' if JA_CONCILIADO else 'CRU'} — ajustando leitura")

C_REF    = _find_col(df, 'referencia', 6, 'p6')
C_TXT    = _find_col(df, 'texto', 5, 'p6')
C_DATA   = _find_col(df, 'data de lancamento', 10 + COL_OFF, 'p6')
C_VALOR  = _find_col(df, 'valor em moeda da empresa', 17 + COL_OFF, 'p6')
C_EST    = _find_col(df, 'estornado', 10, 'p6')
C_DOCEST = _find_col(df, 'documento de estorno', 11, 'p6')

ref = df.iloc[:, C_REF].astype(str).str.strip()
P = pd.DataFrame({
    'ref':   ref,
    'txt':   df.iloc[:, C_TXT].astype(str).str.strip(),
    'valor': pd.to_numeric(df.iloc[:, C_VALOR], errors='coerce'),
    'data':  pd.to_datetime(df.iloc[:, C_DATA], errors='coerce'),
    'sufixo': ref.str.split('-').str[1].fillna(''),
    'num':   pd.to_numeric(ref.str.split('-').str[0], errors='coerce'),
})
_e1 = df.iloc[:, C_EST].astype(str).str.strip().str.upper() == 'X'
_e2 = df.iloc[:, C_DOCEST].astype(str).str.strip().str.upper() == 'X'
P['estornado'] = _e1 | _e2
P['linha'] = P.index + LINHA_OFF
P = P.dropna(subset=['valor', 'num'])
P['cent'] = P['valor'].round(2)

anulacoes = P[P['sufixo'] == '002'].copy()
print(f'  linhas de anulacao (-002): {len(anulacoes)}')
print(f'  ja estornadas (ignoradas): {int(anulacoes["estornado"].sum())}')

# ---------------------------------------------------------------
# RELATORIO DO ERP (opcional) — anulacao -> nota de origem
# ---------------------------------------------------------------
erp = {}   # NF anulacao -> NF origem
if RELATORIO:
    print('Carregando relatorio do ERP...')
    R = pd.read_excel(RELATORIO, header=0)
    r_anul = _find_col(R, 'numero de nota fiscal eletronica', 31, 'erp')
    r_orig = _find_col(R, 'n nota origem', 57, 'erp')
    tab = pd.DataFrame({
        'anul': pd.to_numeric(R.iloc[:, r_anul], errors='coerce'),
        'orig': pd.to_numeric(R.iloc[:, r_orig], errors='coerce'),
    }).dropna().drop_duplicates()
    for _, x in tab.iterrows():
        erp.setdefault(int(x['anul']), int(x['orig']))
    print(f'  anulacoes mapeadas no ERP: {len(erp)}')
else:
    print('Sem relatorio do ERP — usando somente o metodo de valor unico.')

# ---------------------------------------------------------------
# PAREAMENTO
# ---------------------------------------------------------------
# candidatas a origem: qualquer linha que nao seja anulacao e nao esteja estornada
cands = P[(P['sufixo'] != '002') & (~P['estornado'])]

# quantas anulacoes buscam cada valor (para barrar disputa no metodo de valor)
from collections import Counter
busca_valor = Counter(round(-v, 2) for v in anulacoes[~anulacoes['estornado']]['cent'])
qtd_valor   = Counter(cands['cent'])

pares, pendentes = [], []
consumidas = set()   # linhas de origem ja usadas — cada uma serve a UMA anulacao

for _, a in anulacoes.iterrows():
    nf = int(a['num'])
    if a['estornado']:
        pendentes.append((a, None, 'anulacao ja estornada'))
        continue

    alvo = round(-a['cent'], 2)
    escolha, metodo = None, None

    # --- metodo 1: ERP ---
    if nf in erp:
        origem_nf = erp[nf]
        c = cands[(cands['num'] == origem_nf) & (cands['cent'] == alvo)]
        livres = c[~c['linha'].isin(consumidas)]
        if len(livres) > 0:
            escolha, metodo = livres.iloc[0], 'ERP'
        elif len(c) > 0:
            pendentes.append((a, origem_nf, 'origem ja consumida por outra anulacao'))
            continue
        else:
            existe = P[(P['num'] == origem_nf) & (P['cent'] == alvo)]
            motivo = 'origem estornada' if len(existe) else 'origem nao esta na Planilha6'
            pendentes.append((a, origem_nf, motivo))
            continue

    # --- metodo 2: valor unico ---
    else:
        if qtd_valor.get(alvo, 0) == 1 and busca_valor[alvo] == 1:
            c = cands[cands['cent'] == alvo]
            livres = c[~c['linha'].isin(consumidas)]
            if len(livres) > 0:
                escolha, metodo = livres.iloc[0], 'valor unico'
        if escolha is None:
            n = qtd_valor.get(alvo, 0)
            motivo = ('nenhuma linha com o valor invertido' if n == 0
                      else f'ambiguo — {n} candidatas com o mesmo valor')
            pendentes.append((a, None, motivo))
            continue

    consumidas.add(escolha['linha'])
    pares.append({
        'NF Anulacao': nf, 'Ref Anulacao': a['ref'], 'Linha Excel Anul': int(a['linha']),
        'Data Anulacao': a['data'], 'Valor Anulacao': round(a['valor'], 2),
        'NF Origem': int(escolha['num']), 'Ref Origem': escolha['ref'],
        'Linha Excel Orig': int(escolha['linha']), 'Data Origem': escolha['data'],
        'Valor Origem': round(escolha['valor'], 2),
        'Diferenca': round(a['valor'] + escolha['valor'], 2),
        'Status': 'PAR CONFIRMADO', 'Metodo': metodo,
    })

df_pares = pd.DataFrame(pares)
saldo = round(df_pares[['Valor Anulacao', 'Valor Origem']].sum().sum(), 2) if len(df_pares) else 0.0

linhas_pend = [{
    'NF Anulacao': int(a['num']), 'Ref Anulacao': a['ref'], 'Linha Excel Anul': int(a['linha']),
    'Data Anulacao': a['data'], 'Valor Anulacao': round(a['valor'], 2),
    'NF Origem': o, 'Ref Origem': None, 'Linha Excel Orig': None, 'Data Origem': None,
    'Valor Origem': None, 'Diferenca': None, 'Status': f'PENDENTE — {m}', 'Metodo': '-',
} for a, o, m in pendentes]

resultado = pd.concat([df_pares, pd.DataFrame(linhas_pend)], ignore_index=True) \
    if linhas_pend else df_pares

# ---------------------------------------------------------------
# PINTURA + ABA DE RELATORIO
# ---------------------------------------------------------------
base, ext = os.path.splitext(ARQUIVO)
saida = f'{base}_ANULACOES{ext}'
n = 2
while os.path.exists(saida):
    saida = f'{base}_ANULACOES_v{n}{ext}'
    n += 1

print('Pintando as anulacoes...')
wb = load_workbook(ARQUIVO)
ws = wb[P6_NOME]
roxo = PatternFill('solid', fgColor=ROXO_HEX)
alvo_linhas = set(df_pares['Linha Excel Anul']) | set(df_pares['Linha Excel Orig']) if len(df_pares) else set()
for r in sorted(alvo_linhas):
    for c in range(1, ws.max_column + 1):
        ws.cell(row=r, column=c).fill = roxo

# confere o saldo lendo direto das celulas pintadas (prova real)
col_valor_excel = C_VALOR + 1
conf = 0.0
for r in alvo_linhas:
    v = ws.cell(row=r, column=col_valor_excel).value
    if isinstance(v, (int, float)):
        conf += v
conf = round(conf, 2)

if 'Anulacoes' in wb.sheetnames:
    del wb['Anulacoes']
aba = wb.create_sheet('Anulacoes')
cols = list(resultado.columns)
aba.append(cols)
for _, x in resultado.iterrows():
    aba.append([None if pd.isna(v) else (v.to_pydatetime() if isinstance(v, pd.Timestamp) else v)
                for v in x.tolist()])

hf = PatternFill('solid', fgColor='1F4E79')
hfont = Font(bold=True, color='FFFFFF')
verm = PatternFill('solid', fgColor='FFC7CE')
for c in range(1, aba.max_column + 1):
    aba.cell(row=1, column=c).fill = hf
    aba.cell(row=1, column=c).font = hfont
    aba.column_dimensions[get_column_letter(c)].width = 18
aba.freeze_panes = 'A2'
i_st = cols.index('Status') + 1
for r in range(2, aba.max_row + 1):
    for nome in ('Data Anulacao', 'Data Origem'):
        aba.cell(row=r, column=cols.index(nome) + 1).number_format = 'DD/MM/YYYY'
    for nome in ('Valor Anulacao', 'Valor Origem', 'Diferenca'):
        aba.cell(row=r, column=cols.index(nome) + 1).number_format = '#,##0.00'
    fill = roxo if str(aba.cell(row=r, column=i_st).value).startswith('PAR') else verm
    for c in range(1, aba.max_column + 1):
        aba.cell(row=r, column=c).fill = fill

wb.save(saida)

# ---------------------------------------------------------------
# RESUMO
# ---------------------------------------------------------------
print()
print('=' * 62)
print('CONCILIACAO DE ANULACOES')
print('=' * 62)
print(f'  Anulacoes (-002) na planilha : {len(anulacoes)}')
print(f'  Pares confirmados            : {len(df_pares)}   ({len(alvo_linhas)} linhas pintadas)')
if len(df_pares):
    for m, q in df_pares['Metodo'].value_counts().items():
        print(f'      via {m:<12}: {q}')
print(f'  Pendentes de revisao manual  : {len(linhas_pend)}')
for x in linhas_pend:
    print(f"      NF {x['NF Anulacao']}: {x['Status'][10:]}")
print()
print(f'  SALDO DAS LINHAS ROXAS       : R$ {conf:,.2f}')
if abs(conf) <= 0.05:
    print('  [OK] Filtrar pela cor ROXA soma R$0 — pareamento consistente.')
else:
    print('  [ATENCAO] O saldo deveria ser R$0. Revise os pares acima.')
print()
print(f'Arquivo salvo: {saida}')
print("Aba 'Anulacoes' criada com o detalhe de cada par (roxo) e pendencia (vermelho).")
